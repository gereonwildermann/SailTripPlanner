import os
import pandas as pd
from openpyxl import load_workbook

# Dateipfad
excel_path = 'Einkaufsliste.xlsx'

def create_default_dataframes():
    """Erstellt Standard-Datenframes mit leeren Werten oder Default-Inhalten."""
    # Daten für Frühstück, Mittagessen, Abendessen, Getränke, Gewürze und Kräuter, Snacks
    fruehstueck = pd.DataFrame({
        'Gericht': [],
        'Zutat': [],
        'Menge': [],
        'Einheit': [],
        'Notizen': [],
        'Kategorie': []
    })

    mittagessen = pd.DataFrame({
        'Gericht': [],
        'Zutat': [],
        'Menge': [],
        'Einheit': [],
        'Notizen': [],
        'Kategorie': []
    })

    abendessen = pd.DataFrame({
        'Gericht': [],
        'Zutat': [],
        'Menge': [],
        'Einheit': [],
        'Notizen': [],
        'Kategorie': []
    })

    getraenke = pd.DataFrame({
        'Getränk': ['Bier', 'Wasser', 'Gin Tonic', 'Gin Tonic'],
        'Zutat': ['Bier', 'Wasser', 'Gin', 'Tonic Water'],
        'Menge': ['', '', '', ''],
        'Einheit': ['Liter', 'Liter', '', ''],
        'Notizen': ['','still und sprudelnd', '', ''],
        'Kategorie': ['Alkoholische Getränke','Alkoholfreie Getränke', 'Alkoholische Getränke', 'Alkoholfreie Getränke']
    })

    gewuerze_und_kraeuter = pd.DataFrame({
        'Gewürz/Kraut/Öl': ['Salz', 'Pfeffer', 'Olivenöl', 'Balsamico'],
        'Menge': ['', '', '',''],
        'Einheit': ['Gramm', 'Gramm', 'Flaschen', 'Flaschen'],
        'Notizen': ['Feines Salz', '', '', ''],
        'Kategorie': ['Gewürze', 'Gewürze', 'Öle', 'Gewürze']
    })

    snacks = pd.DataFrame({
        'Snack': ['Chips', 'Nüsse', 'Kekse'],
        'Menge': ['', '', ''],
        'Einheit': ['', '', ''],
        'Notizen': ['', '', ''],
        'Kategorie': ['Snacks', 'Snacks', 'Snacks']
    })
    
    return fruehstueck, mittagessen, abendessen, getraenke, gewuerze_und_kraeuter, snacks

def create_einkaufsliste(*dfs):
    """Erstellt die Einkaufsliste basierend auf den übergebenen DataFrames ohne Gruppierung."""
    # DataFrames zusammenführen
    all_data = pd.concat(dfs, ignore_index=True)
    
    # Debugging-Ausgabe: Überprüfen der zusammengeführten Daten
    print("Kombinierte DataFrames:")
    print(all_data.head())
    
    # Spalten konsistent umbenennen
    all_data['Artikel'] = all_data.get('Zutat', '') \
                        .fillna(all_data.get('Gewürz/Kraut/Öl', '')) \
                        .fillna(all_data.get('Snack', ''))
    
    # Entfernen der alten Spalten
    all_data = all_data[['Artikel', 'Menge', 'Einheit', 'Kategorie', 'Notizen']]
    
    # Debugging-Ausgabe: Überprüfen der Spaltennamen nach der Umbenennung
    print("DataFrame nach Umbenennen der Spalten:")
    print(all_data.head())
    
    # Bereinigen der DataFrame für die Ausgabe
    all_data['Menge'] = pd.to_numeric(all_data['Menge'], errors='coerce').fillna(0)
    
    # Debugging-Ausgabe: Überprüfen der Daten vor der Ausgabe
    print("DataFrame vor der Ausgabe:")
    print(all_data.head())
    
    # Sortieren nach Kategorie
    einkaufsliste = all_data.sort_values(by='Kategorie')
    
    # Debugging-Ausgabe: Überprüfen der finalen Einkaufsliste
    print("Einkaufsliste:")
    print(einkaufsliste.head())
    
    return einkaufsliste

def update_excel_file():
    """Aktualisiert die Excel-Datei, je nachdem ob sie bereits existiert oder nicht."""
    if not os.path.exists(excel_path):
        # Datei erstellen und Standardinhalte hinzufügen
        fruehstueck, mittagessen, abendessen, getraenke, gewuerze_und_kraeuter, snacks = create_default_dataframes()
        einkaufsliste = create_einkaufsliste(fruehstueck, mittagessen, abendessen, getraenke, gewuerze_und_kraeuter, snacks)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            fruehstueck.to_excel(writer, sheet_name='Frühstück', index=False)
            mittagessen.to_excel(writer, sheet_name='Mittagessen', index=False)
            abendessen.to_excel(writer, sheet_name='Abendessen', index=False)
            getraenke.to_excel(writer, sheet_name='Getränke', index=False)
            gewuerze_und_kraeuter.to_excel(writer, sheet_name='Gewürze, Kräuter und Öle', index=False)
            snacks.to_excel(writer, sheet_name='Snacks', index=False)
            einkaufsliste.to_excel(writer, sheet_name='Einkaufsliste', index=False)
    
    else:
        # Datei einlesen und vorhandene Daten verarbeiten
        workbook = load_workbook(excel_path)
        sheets = workbook.sheetnames
        
        # Laden der Daten aus den existierenden Blättern
        fruehstueck = pd.read_excel(excel_path, sheet_name='Frühstück')
        mittagessen = pd.read_excel(excel_path, sheet_name='Mittagessen')
        abendessen = pd.read_excel(excel_path, sheet_name='Abendessen')
        getraenke = pd.read_excel(excel_path, sheet_name='Getränke')
        gewuerze_und_kraeuter = pd.read_excel(excel_path, sheet_name='Gewürze, Kräuter und Öle')
        snacks = pd.read_excel(excel_path, sheet_name='Snacks')
        
        einkaufsliste = create_einkaufsliste(fruehstueck, mittagessen, abendessen, getraenke, gewuerze_und_kraeuter, snacks)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            for sheet in sheets:
                if sheet != 'Einkaufsliste':
                    df = pd.read_excel(excel_path, sheet_name=sheet)
                    df.to_excel(writer, sheet_name=sheet, index=False)
            einkaufsliste.to_excel(writer, sheet_name='Einkaufsliste', index=False)
    
    # Excel-Datei speichern
    # Keine Dropdown-Listen hinzufügen
    workbook = load_workbook(excel_path)
    workbook.save(excel_path)

# Update Excel-Datei
update_excel_file()
